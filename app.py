from __future__ import annotations

from flask import Flask, render_template, request, redirect, url_for, session, flash, Response, abort,jsonify
from datetime import datetime

import csv
import io

from database import init_db
from models import (
    add_client,
    create_user,
    delete_client,
    add_task,
    delete_task,
    get_client_by_id,
    get_clients_for_user,
    get_task_by_id,
    get_tasks_for_client,
    get_user_by_email,
    is_logged_in,
    logout_user,
    update_client,
    update_task,
    verify_password,
    get_clients_and_tasks_for_user,
)

from services.ai_service import (
    OllamaError,
    OllamaNotRunningError,
    OllamaModelNotFoundError,
    OllamaTimeoutError,
    ollama_generate,
    generate_insights,
    generate_daily_summary,
    generate_weekly_report,
    ask_ai,
    explain_dashboard,
    sql_helper,
    generate_recommendations,
)

from services.chat_history_utils import (
    clear_chat_history_for_user,
    get_chat_history_for_user,
)


app = Flask(__name__)
app.secret_key = "change-this-secret-key"  

# Initialize database tables
init_db()


@app.route("/")
def home():
    """Redirect user depending on login state."""
    if is_logged_in(session):
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/signup", methods=["GET", "POST"])
def signup():
    """User registration."""
    if is_logged_in(session):
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        if not username or not email or not password:
            flash("All fields are required.")
            return redirect(url_for("signup"))

        user = create_user(username=username, email=email, password=password)
        if user is None:
            flash("A user with this email already exists.")
            return redirect(url_for("signup"))

        # Log user in immediately after signup
        session["user_id"] = user["id"]
        session["email"] = user["email"]
        return redirect(url_for("dashboard"))

    return render_template("signup.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    """User login."""
    if is_logged_in(session):
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        user = get_user_by_email(email)
        if user is None:
            flash("Invalid email or password.")
            return redirect(url_for("login"))

        if not verify_password(user, password):
            flash("Invalid email or password.")
            return redirect(url_for("login"))

        session["user_id"] = user["id"]
        session["email"] = user["email"]
        session["username"] = user.get("username")
        return redirect(url_for("dashboard"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    """Log user out."""
    logout_user(session)
    return redirect(url_for("login"))


@app.route("/dashboard")
def dashboard():

    if not is_logged_in(session):
        return redirect(url_for("login"))

    user_id = session.get("user_id")

    from database import get_connection

    conn = get_connection()
    cur = conn.cursor()

    total_clients = len(get_clients_for_user(user_id=user_id))

    cur.execute(
        """
        SELECT COUNT(*) as cnt
        FROM tasks t
        JOIN clients c ON c.id = t.client_id
        WHERE c.user_id = ?
        """,
        (user_id,),
    )
    total_tasks = cur.fetchone()[0] or 0

    cur.execute(
        """
        SELECT t.status, COUNT(*) as cnt
        FROM tasks t
        JOIN clients c ON c.id = t.client_id
        WHERE c.user_id = ?
        GROUP BY t.status
        """,
        (user_id,),
    )
    status_rows = cur.fetchall() or []
    status_counts = {row[0]: (row[1] or 0) for row in status_rows}

    pending_tasks = status_counts.get("Pending", 0)
    in_progress_tasks = status_counts.get("In Progress", 0)
    completed_tasks = status_counts.get("Completed", 0)

    cur.execute(
        """
        SELECT COUNT(*) as cnt
        FROM tasks t
        JOIN clients c ON c.id = t.client_id
        WHERE c.user_id = ? AND t.priority = 'High'
        """,
        (user_id,),
    )
    high_priority_tasks = cur.fetchone()[0] or 0

    cur.execute(
        """
        SELECT strftime('%Y-%m', created_at) as ym, COUNT(*) as cnt
        FROM clients
        WHERE user_id = ?
        GROUP BY ym
        ORDER BY ym ASC
        """,
        (user_id,),
    )
    month_rows = cur.fetchall() or []
    client_months = [r[0] for r in month_rows if r[0] is not None]
    client_month_counts = [r[1] or 0 for r in month_rows]

    pie_labels = ["Pending", "In Progress", "Completed"]
    pie_values = [
        status_counts.get("Pending", 0),
        status_counts.get("In Progress", 0),
        status_counts.get("Completed", 0),
    ]

    cur.execute(
        """
        SELECT t.priority, COUNT(*) as cnt
        FROM tasks t
        JOIN clients c ON c.id = t.client_id
        WHERE c.user_id = ?
        GROUP BY t.priority
        """,
        (user_id,),
    )
    priority_rows = cur.fetchall() or []
    priority_counts = {row[0]: (row[1] or 0) for row in priority_rows}

    doughnut_labels = ["High", "Medium", "Low"]
    doughnut_values = [
        priority_counts.get("High", 0),
        priority_counts.get("Medium", 0),
        priority_counts.get("Low", 0),
    ]

    cur.execute(
        """
        SELECT id, full_name, email, created_at
        FROM clients
        WHERE user_id = ?
        ORDER BY created_at DESC
        LIMIT 5
        """,
        (user_id,),
    )
    recent_clients_rows = cur.fetchall() or []
    recent_clients = [
        {"id": r[0], "full_name": r[1], "email": r[2], "created_at": r[3]}
        for r in recent_clients_rows
    ]

    cur.execute(
        """
        SELECT t.id, t.title, t.status, t.priority, t.created_at,
               c.full_name as client_name
        FROM tasks t
        JOIN clients c ON c.id = t.client_id
        WHERE c.user_id = ?
        ORDER BY t.created_at DESC
        LIMIT 5
        """,
        (user_id,),
    )
    recent_tasks_rows = cur.fetchall() or []
    recent_tasks = [
        {
            "id": r[0],
            "title": r[1],
            "status": r[2],
            "priority": r[3],
            "created_at": r[4],
            "client_name": r[5],
        }
        for r in recent_tasks_rows
    ]

    conn.close()

    return render_template(
        "dashboard.html",
        today_date=datetime.now().strftime('%Y-%m-%d'),
        total_clients=total_clients,
        total_tasks=total_tasks,
        pending_tasks=pending_tasks,
        in_progress_tasks=in_progress_tasks,
        completed_tasks=completed_tasks,
        high_priority_tasks=high_priority_tasks,
        client_months=client_months,
        client_month_counts=client_month_counts,
        pie_labels=pie_labels,
        pie_values=pie_values,
        doughnut_labels=doughnut_labels,
        doughnut_values=doughnut_values,
        recent_clients=recent_clients,
        recent_tasks=recent_tasks,
    )


@app.route("/clients", methods=["GET"])
def clients_list():
    """List clients for the logged-in user (Step-2 search/filter/sort)."""
    if not is_logged_in(session):
        return redirect(url_for("login"))

    user_id = session.get("user_id")

    
    q = request.args.get("q", "").strip() or None  # full_name/company/email
    company_filter = request.args.get("company", "").strip() or None
    sort = request.args.get("sort", "newest").strip().lower()
    if sort not in ("newest", "oldest", "alpha_az", "alpha_za"):
        sort = "newest"

    clients = get_clients_for_user(
        user_id=user_id,
        q=q,
        company_filter=company_filter,
        sort=sort,
    )

    return render_template(
        "clients.html",
        clients=clients,
        search=q,
        company_filter=company_filter,
        sort=sort,
        editing_client=None,
        form_data=None,
    )


@app.route("/clients/add", methods=["POST"])
def clients_add():
    """Create a new client (Step-2)."""
    if not is_logged_in(session):
        return redirect(url_for("login"))

    user_id = session.get("user_id")

    full_name = request.form.get("full_name", "")
    email = request.form.get("email", "")
    phone = request.form.get("phone", "")
    company = request.form.get("company", "")
    address = request.form.get("address", "")

    _, error = add_client(
        user_id=user_id,
        full_name=full_name,
        email=email,
        phone=phone,
        company=company,
        address=address,
    )

    # Keep search params if validation fails
    q = request.args.get("q", "").strip() or None
    company_filter = request.args.get("company", "").strip() or None
    sort = request.args.get("sort", "newest").strip().lower()
    if sort not in ("newest", "oldest", "alpha_az", "alpha_za"):
        sort = "newest"

    if error:
        flash(error)
        clients = get_clients_for_user(
            user_id=user_id,
            q=q,
            company_filter=company_filter,
            sort=sort,
        )
        return render_template(
            "clients.html",
            clients=clients,
            search=q,
            company_filter=company_filter,
            sort=sort,
            editing_client=None,
            form_data={
                "full_name": full_name,
                "email": email,
                "phone": phone,
                "company": company,
                "address": address,
            },
        )

    return redirect(url_for("clients_list"))


@app.route("/clients/edit/<int:client_id>", methods=["GET"])
def clients_edit_get(client_id: int):
    """Load a client into the edit form (ownership enforced by model)."""
    if not is_logged_in(session):
        return redirect(url_for("login"))

    user_id = session.get("user_id")
    client = get_client_by_id(user_id=user_id, client_id=client_id)

    if not client:
        flash("Client not found (or you don't have access).")
        return redirect(url_for("clients_list"))

    clients = get_clients_for_user(user_id=user_id)

    return render_template(
        "clients.html",
        clients=clients,
        search=None,
        company_filter=None,
        sort="newest",
        editing_client=client,
        form_data={
            "full_name": client.get("full_name"),
            "email": client.get("email") or "",
            "phone": client.get("phone") or "",
            "company": client.get("company") or "",
            "address": client.get("address") or "",
        },
    )


@app.route("/clients/edit/<int:client_id>", methods=["POST"])
def clients_edit_post(client_id: int):
    """Update a client (ownership enforced by model)."""
    if not is_logged_in(session):
        return redirect(url_for("login"))

    user_id = session.get("user_id")

    full_name = request.form.get("full_name", "")
    email = request.form.get("email", "")
    phone = request.form.get("phone", "")
    company = request.form.get("company", "")
    address = request.form.get("address", "")

    _, error = update_client(
        user_id=user_id,
        client_id=client_id,
        full_name=full_name,
        email=email,
        phone=phone,
        company=company,
        address=address,
    )

    if error:
        flash(error)
        client = get_client_by_id(user_id=user_id, client_id=client_id)
        clients = get_clients_for_user(user_id=user_id)
        return render_template(
            "clients.html",
            clients=clients,
            search=None,
            company_filter=None,
            sort="newest",
            editing_client=client,
            form_data={
                "full_name": full_name,
                "email": email,
                "phone": phone,
                "company": company,
                "address": address,
            },
        )

    return redirect(url_for("clients_list"))


@app.route("/clients/delete/<int:client_id>", methods=["POST"])
def clients_delete(client_id: int):
    """Delete a client (ownership enforced by model)."""
    if not is_logged_in(session):
        return redirect(url_for("login"))

    user_id = session.get("user_id")
    ok = delete_client(user_id=user_id, client_id=client_id)

    if not ok:
        flash("Client not found (or you don't have access).")

    return redirect(url_for("clients_list"))


@app.route("/tasks/<int:client_id>", methods=["GET"])
def tasks_list(client_id: int):
    """List tasks for the selected client."""
    if not is_logged_in(session):
        return redirect(url_for("login"))

    user_id = session.get("user_id")

    selected_client = get_client_by_id(user_id=user_id, client_id=client_id)
    if not selected_client:
        flash("Client not found (or you don't have access).")
        return redirect(url_for("clients_list"))

    status = request.args.get("status", "").strip() or None
    priority = request.args.get("priority", "").strip() or None
    due_date = request.args.get("due_date", "").strip() or None

    tasks = get_tasks_for_client(
        user_id=user_id,
        client_id=client_id,
        status=status,
        priority=priority,
        due_date=due_date,
    )

    return render_template(
        "tasks.html",
        selected_client=selected_client,
        selected_client_id=client_id,
        tasks=tasks,
        editing_task=None,
        status=status,
        priority=priority,
        due_date=due_date,
    )


@app.route("/tasks/<int:client_id>/add", methods=["POST"])
def tasks_add(client_id: int):
    """Create a new task for the selected client."""
    if not is_logged_in(session):
        return redirect(url_for("login"))

    user_id = session.get("user_id")

    title = request.form.get("title", "")
    description = request.form.get("description", "")
    priority = request.form.get("priority", "")
    status = request.form.get("status", "")
    due_date = request.form.get("due_date", "").strip() or None

    _, error = add_task(
        user_id=user_id,
        client_id=client_id,
        title=title,
        description=description,
        priority=priority,
        status=status,
        due_date=due_date,
    )

    if error:
        flash(error)

    return redirect(url_for("tasks_list", client_id=client_id))


@app.route("/tasks/<int:client_id>/edit/<int:task_id>", methods=["GET"])
def tasks_edit_get(task_id: int, client_id: int):
    """Load a task into the edit form."""
    if not is_logged_in(session):
        return redirect(url_for("login"))

    user_id = session.get("user_id")

    selected_client = get_client_by_id(user_id=user_id, client_id=client_id)
    if not selected_client:
        flash("Client not found (or you don't have access).")
        return redirect(url_for("clients_list"))

    editing_task = get_task_by_id(user_id=user_id, task_id=task_id)
    if not editing_task:
        flash("Task not found (or you don't have access).")
        return redirect(url_for("tasks_list", client_id=client_id))

    return render_template(
        "tasks.html",
        selected_client=selected_client,
        selected_client_id=client_id,
        tasks=get_tasks_for_client(user_id=user_id, client_id=client_id),
        editing_task=editing_task,
        status=None,
        priority=None,
        due_date=None,
    )


@app.route("/tasks/<int:client_id>/edit/<int:task_id>", methods=["POST"])
def tasks_edit_post(task_id: int, client_id: int):
    """Update a task."""
    if not is_logged_in(session):
        return redirect(url_for("login"))

    user_id = session.get("user_id")

    title = request.form.get("title", "")
    description = request.form.get("description", "")
    priority = request.form.get("priority", "")
    status = request.form.get("status", "")
    due_date = request.form.get("due_date", "").strip() or None

    _, error = update_task(
        user_id=user_id,
        task_id=task_id,
        title=title,
        description=description,
        priority=priority,
        status=status,
        due_date=due_date,
    )

    if error:
        flash(error)
        selected_client = get_client_by_id(user_id=user_id, client_id=client_id)
        editing_task = get_task_by_id(user_id=user_id, task_id=task_id)
        return render_template(
            "tasks.html",
            selected_client=selected_client,
            selected_client_id=client_id,
            tasks=get_tasks_for_client(user_id=user_id, client_id=client_id),
            editing_task=editing_task,
            status=None,
            priority=None,
            due_date=None,
        )

    flash("Task updated successfully.")
    return redirect(url_for("tasks_list", client_id=client_id))


@app.route("/tasks/delete/<int:task_id>", methods=["POST"])
def tasks_delete(task_id: int):
    """Delete a task (with ownership enforced by lookup)."""
    if not is_logged_in(session):
        return redirect(url_for("login"))

    user_id = session.get("user_id")

    existing = get_task_by_id(user_id=user_id, task_id=task_id)
    if not existing:
        flash("Task not found (or you don't have access).")
        return redirect(url_for("dashboard"))

    ok = delete_task(user_id=user_id, task_id=task_id)
    if not ok:
        flash("Unable to delete task.")
    else:
        flash("Task deleted successfully.")

    return redirect(url_for("tasks_list", client_id=existing["client_id"]))


@app.route("/export/csv")
def export_csv():
    if not is_logged_in(session):
        return redirect(url_for("login"))

    rows = get_clients_and_tasks_for_user(session.get("user_id"))

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(
        [
            "Client Name",
            "Email",
            "Phone",
            "Company",
            "Address",
            "Task Title",
            "Task Description",
            "Priority",
            "Status",
            "Due Date",
            "Client Created At",
        ]
    )

    for r in rows:
        writer.writerow(
            [
                r.get("client_name") or "-",
                r.get("client_email") or "-",
                r.get("client_phone") or "-",
                r.get("client_company") or "-",
                r.get("client_address") or "-",
                r.get("task_title") or "-",
                r.get("task_description") or "-",
                r.get("task_priority") or "-",
                r.get("task_status") or "-",
                r.get("task_due_date") or "-",
                r.get("client_created_at") or "-",
            ]
        )

    output.seek(0)

    return Response(
        output.getvalue().encode("utf-8-sig"),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=export-CSV.csv"},
    )


@app.route("/export/excel")
def export_excel():
    if not is_logged_in(session):
        return redirect(url_for("login"))

    from io import BytesIO

    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    rows = get_clients_and_tasks_for_user(session.get("user_id"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"

    headers = [
        "Client Name",
        "Email",
        "Phone",
        "Company",
        "Address",
        "Task Title",
        "Task Description",
        "Priority",
        "Status",
        "Due Date",
        "Client Created At",
    ]

    header_font = Font(bold=True)

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = header_font

    for r in rows:
        ws.append(
            [
                r.get("client_name") or "",
                r.get("client_email") or "",
                r.get("client_phone") or "",
                r.get("client_company") or "",
                r.get("client_address") or "",
                r.get("task_title") or "",
                r.get("task_description") or "",
                r.get("task_priority") or "",
                r.get("task_status") or "",
                r.get("task_due_date") or "",
                r.get("client_created_at") or "",
            ]
        )

    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1):
            for c in cell:
                if c.value is None:
                    continue
                max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 55)

    ws.freeze_panes = "A2"

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return Response(
        stream.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=export-excel.xlsx"},
    )


@app.route("/export/pdf")
def export_pdf():
    if not is_logged_in(session):
        return redirect(url_for("login"))

    from io import BytesIO

    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    rows = get_clients_and_tasks_for_user(session.get("user_id"))

    client_names = set()
    total_tasks = 0
    for r in rows:
        client_names.add(r.get("client_name"))
        if r.get("task_title"):
            total_tasks += 1
    total_clients = len(client_names)

    generation_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    width, height = letter

    left_margin = 40
    top_margin = height - 40
    y = top_margin

    def draw_header():
        nonlocal y
        c.setFont("Helvetica-Bold", 16)
        c.drawString(left_margin, y, "Client Management Report")
        y -= 20

        c.setFont("Helvetica", 10)
        c.drawString(left_margin, y, f"Generation date: {generation_date}")
        y -= 14
        c.drawString(left_margin, y, f"Total clients: {total_clients}")
        y -= 14
        c.drawString(left_margin, y, f"Total tasks: {total_tasks}")
        y -= 20

    def draw_table_header():
        nonlocal y
        c.setFont("Helvetica-Bold", 9)
        headers = ["Client", "Company", "Email", "Phone", "Task", "Priority", "Status", "Due Date"]
        col_widths = [90, 70, 120, 70, 100, 55, 70, 65]
        x = left_margin
        for h, w in zip(headers, col_widths):
            c.drawString(x, y, h)
            x += w
        y -= 12

    def new_page():
        nonlocal y
        c.showPage()
        y = top_margin
        draw_header()
        draw_table_header()

    draw_header()
    draw_table_header()

    c.setFont("Helvetica", 8)
    col_widths = [90, 70, 120, 70, 100, 55, 70, 65]

    for r in rows:
        x = left_margin
        row_values = [
            r.get("client_name") or "",
            r.get("client_company") or "",
            r.get("client_email") or "",
            r.get("client_phone") or "",
            r.get("task_title") or "",
            r.get("task_priority") or "",
            r.get("task_status") or "",
            r.get("task_due_date") or "",
        ]
        for val, w in zip(row_values, col_widths):
            text = str(val)
            if len(text) > 28:
                text = text[:27] + "…"
            c.drawString(x, y, text)
            x += w

        y -= 12
        if y < 60:
            new_page()

    c.save()
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": "attachment; filename=export-pdf.pdf"},
    )


@app.route("/export/legacy")
def export_legacy():
    abort(404)


@app.route("/assistant")
def assistant_page():
    if not is_logged_in(session):
        return redirect(url_for("login"))

    from services.ai_service import DEFAULT_MODEL

    return render_template("assistant.html", model=DEFAULT_MODEL)


@app.route("/api/assistant/history", methods=["GET"])
def api_assistant_history():
    if not is_logged_in(session):
        return {"error": "Unauthorized"}, 401

    user_id = session.get("user_id")
    messages = get_chat_history_for_user(user_id=user_id, limit=20)
    return {"messages": messages}


@app.route("/api/assistant/clear", methods=["POST"])
def api_assistant_clear():
    
    if not is_logged_in(session):
        return {"error": "Unauthorized"}, 401

    user_id = session.get("user_id")
    clear_chat_history_for_user(user_id=user_id)
    return {"ok": True}


@app.route("/api/clear-chat", methods=["POST"])
def api_clear_chat():
        
    if not is_logged_in(session):
        return {"error": "Unauthorized"}, 401

    user_id = session.get("user_id")
    clear_chat_history_for_user(user_id=user_id)
    return {"success": True}


@app.route("/api/ask-ai", methods=["POST"])
def api_ask_ai():

    if not is_logged_in(session):
        return {"error": "Unauthorized"}, 401

    data = request.get_json(silent=True) or {}

    question = (data.get("question") or "").strip()
    mode = (data.get("mode") or "chat").strip()

    user_id = session.get("user_id")

    if not question:
        return {"error": "Question is required."}, 400

    try:
        result = ask_ai(
            question=question,
            user_id=user_id,
            mode=mode
        )

        return {
            "answer": result.get("answer", "")
        }

    except OllamaNotRunningError:
        return {"error": "Ollama is not running."}, 503

    except OllamaModelNotFoundError:
        return {"error": "Model not found in Ollama. Install it with: ollama pull <model>"}, 404

    except OllamaTimeoutError:
        return {"error": "Ollama request timed out."}, 504

    except Exception as e:
        return {"error": f"AI error: {str(e)[:160]}"}, 500
    
    

def register_assistant_history_routes(app, is_logged_in_func, clear_func, get_func):
    @app.route('/api/assistant/history', methods=['GET'])
    def assistant_history():
        if not is_logged_in_func(session):
            return {'error': 'Unauthorized'}, 401

        user_id = session.get('user_id')
        messages = get_func(user_id=user_id)
        return jsonify({'messages': messages})

    @app.route('/api/assistant/clear', methods=['POST'])
    def assistant_clear():
        if not is_logged_in_func(session):
            return {'error': 'Unauthorized'}, 401

        user_id = session.get('user_id')
        clear_func(user_id=user_id)
        return jsonify({'ok': True})

if __name__ == "__main__":
    init_db()
    app.run(debug=True)